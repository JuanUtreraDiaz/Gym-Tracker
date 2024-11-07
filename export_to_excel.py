# export_to_excel.py
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from database import create_connection

def export_to_excel():
    conn = create_connection()
    
    # Exportar cada tabla
    tables = {}
    for table in ["Usuarios", "Ejercicios", "Entrenamientos"]:
        tables[table] = pd.read_sql_query(f"SELECT * FROM {table}", conn)
    
    # Guardar en Excel inicialmente con pandas
    file_path = "data/fitness_app_export.xlsx"
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for table_name, data in tables.items():
            data.to_excel(writer, sheet_name=table_name, index=False)
    
    # Añadir formato de tabla en Excel usando openpyxl
    workbook = load_workbook(file_path)
    for sheet_name in tables.keys():
        worksheet = workbook[sheet_name]
        
        # Definir el rango de la tabla
        last_column = chr(64 + worksheet.max_column)  # Convierte el número de columna a letra
        table_range = f"A1:{last_column}{worksheet.max_row}"
        
        # Crear la tabla con estilo
        table = Table(displayName=sheet_name, ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        table.tableStyleInfo = style
        worksheet.add_table(table)
    
    workbook.save(file_path)
    conn.close()
    print("Base de datos exportada a Excel con formato de tabla")
